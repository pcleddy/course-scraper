#!/usr/bin/env python3
"""
CNM Course Scraper
==================
Scrapes all courses from a Banner class search system and saves them to
CSV, Excel, and a browser-friendly JavaScript bundle.

Usage:
    python cnm_course_scraper.py                  # auto-selects first available term
    python cnm_course_scraper.py --term 202580    # specific term code
    python cnm_course_scraper.py --list-terms     # just list available terms

Requirements:
    pip install requests
    pip install openpyxl   # optional, for Excel output
"""

import requests
import csv
import json
import sys
import time
import argparse
import logging
from pathlib import Path
from datetime import datetime, timezone
from uuid import uuid4

BASE_URL = "https://banner.cnm.edu/StudentRegistrationSsb/ssb"
DEFAULT_SCHOOL_LABEL = "CNM"
DEFAULT_OUTPUT_STEM = "cnm_courses"
DEFAULT_BUNDLE_NAME = "cnm_courses_data.js"

log = logging.getLogger("cnm_scraper")

# Columns to extract from the Banner JSON response
COLUMNS = [
    "term",
    "courseReferenceNumber",
    "subject",
    "subjectDescription",
    "courseNumber",
    "courseTitle",
    "creditHourLow",
    "creditHourHigh",
    "creditHours",
    "maximumEnrollment",
    "enrollment",
    "seatsAvailable",
    "waitCapacity",
    "waitCount",
    "waitAvailable",
    "campusDescription",
    "instructionalMethodDescription",
    "scheduleTypeDescription",
    "sequenceNumber",
    "partOfTermDescription",
    "faculty_names",
    "meeting_days",
    "meeting_times",
    "meeting_building",
    "meeting_room",
    "meeting_type",
    "openSection",
    "linkIdentifier",
    "isSectionLinked",
    "sectionAttributes",
]


class BannerClient:
    """Client for interacting with Ellucian Banner's class search API."""

    def __init__(self, base_url=BASE_URL):
        self.base_url = base_url
        self.session = requests.Session()
        self.session.headers.update({
            "User-Agent": (
                "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
                "AppleWebKit/537.36 (KHTML, like Gecko) "
                "Chrome/120.0.0.0 Safari/537.36"
            ),
            "Accept": "application/json, text/javascript, */*; q=0.01",
            "X-Requested-With": "XMLHttpRequest",
        })
        self._session_id = str(uuid4())

    def initialize(self):
        """Hit the main search page to establish server-side session + cookies."""
        r = self.session.get(
            f"{self.base_url}/classSearch/classSearch",
            timeout=30,
        )
        r.raise_for_status()
        log.debug(f"Session initialized. Cookies: {dict(self.session.cookies)}")
        return r

    def get_terms(self):
        """Fetch the list of available terms (semesters)."""
        r = self.session.get(
            f"{self.base_url}/classSearch/getTerms",
            params={"offset": 1, "max": 50, "searchTerm": ""},
            timeout=30,
        )
        r.raise_for_status()
        terms = r.json()
        log.debug(f"Got {len(terms)} terms")
        return terms  # list of {"code": "202580", "description": "SPRING 2026"}

    def select_term(self, term_code):
        """
        POST the selected term to the server. This is REQUIRED before any
        search will return results — it sets the term in the server-side session.
        """
        r = self.session.post(
            f"{self.base_url}/term/search",
            params={"mode": "search"},
            data={"term": term_code},
            timeout=30,
        )
        r.raise_for_status()
        log.debug(f"Term {term_code} selected. Status: {r.status_code}")
        return r

    def get_subjects(self, term_code):
        """Fetch all subjects for a given term."""
        subjects = []
        offset = 1
        page_size = 500
        while True:
            r = self.session.get(
                f"{self.base_url}/classSearch/get_subject",
                params={
                    "term": term_code,
                    "offset": offset,
                    "max": page_size,
                    "searchTerm": "",
                },
                timeout=30,
            )
            r.raise_for_status()
            page = r.json()
            if not page:
                break
            subjects.extend(page)
            if len(page) < page_size:
                break
            offset += page_size
        log.debug(f"Got {len(subjects)} subjects for term {term_code}")
        return subjects

    def reset_search(self):
        """Reset the server-side search form data."""
        r = self.session.post(
            f"{self.base_url}/classSearch/resetDataForm",
            timeout=30,
        )
        log.debug(f"Search reset. Status: {r.status_code}")
        return r

    def search_courses(self, term_code, subject_code=None, offset=0, page_size=500):
        """
        Execute a course search and return one page of results.
        Returns the raw JSON dict with 'data', 'totalCount', etc.
        """
        params = {
            "txt_term": term_code,
            "startDatepicker": "",
            "endDatepicker": "",
            "uniqueSessionId": self._session_id,
            "pageOffset": offset,
            "pageMaxSize": page_size,
            "sortColumn": "subjectDescription",
            "sortDirection": "asc",
        }
        if subject_code:
            params["txt_subject"] = subject_code

        r = self.session.get(
            f"{self.base_url}/searchResults/searchResults",
            params=params,
            timeout=60,
        )
        r.raise_for_status()
        data = r.json()
        log.debug(
            f"Search subject={subject_code} offset={offset}: "
            f"totalCount={data.get('totalCount', '?')}, "
            f"rows={len(data.get('data', []) or [])}"
        )
        return data


def extract_row(section):
    """Extract a flat dict from a Banner section JSON object."""
    # Faculty names
    faculty = section.get("faculty", []) or []
    faculty_names = "; ".join(
        f.get("displayName", "TBA") for f in faculty
    ) if faculty else "TBA"

    # Meeting times
    meetings = section.get("meetingsFaculty", []) or []
    meeting_days_list = []
    meeting_times_list = []
    meeting_buildings = []
    meeting_rooms = []
    meeting_types = []

    for m in meetings:
        mf = m.get("meetingTime", {}) or {}

        # Days
        days = ""
        for d in ["monday", "tuesday", "wednesday", "thursday", "friday", "saturday", "sunday"]:
            if mf.get(d):
                days += d[:2].capitalize()
        meeting_days_list.append(days or "TBA")

        # Times
        begin = mf.get("beginTime", "")
        end = mf.get("endTime", "")
        if begin and end:
            meeting_times_list.append(f"{begin[:2]}:{begin[2:]}-{end[:2]}:{end[2:]}")
        else:
            meeting_times_list.append("TBA")

        meeting_buildings.append(mf.get("buildingDescription", "") or "")
        meeting_rooms.append(mf.get("room", "") or "")
        meeting_types.append(mf.get("meetingTypeDescription", "") or "")

    # Section attributes
    attrs = section.get("sectionAttributes", []) or []
    attr_str = "; ".join(a.get("description", "") for a in attrs)

    return {
        "term": section.get("term", ""),
        "courseReferenceNumber": section.get("courseReferenceNumber", ""),
        "subject": section.get("subject", ""),
        "subjectDescription": section.get("subjectDescription", ""),
        "courseNumber": section.get("courseNumber", ""),
        "courseTitle": section.get("courseTitle", ""),
        "creditHourLow": section.get("creditHourLow", ""),
        "creditHourHigh": section.get("creditHourHigh", ""),
        "creditHours": section.get("creditHours", ""),
        "maximumEnrollment": section.get("maximumEnrollment", ""),
        "enrollment": section.get("enrollment", ""),
        "seatsAvailable": section.get("seatsAvailable", ""),
        "waitCapacity": section.get("waitCapacity", ""),
        "waitCount": section.get("waitCount", ""),
        "waitAvailable": section.get("waitAvailable", ""),
        "campusDescription": section.get("campusDescription", ""),
        "instructionalMethodDescription": section.get("instructionalMethodDescription", ""),
        "scheduleTypeDescription": section.get("scheduleTypeDescription", ""),
        "sequenceNumber": section.get("sequenceNumber", ""),
        "partOfTermDescription": section.get("partOfTermDescription", ""),
        "faculty_names": faculty_names,
        "meeting_days": " | ".join(meeting_days_list),
        "meeting_times": " | ".join(meeting_times_list),
        "meeting_building": " | ".join(meeting_buildings),
        "meeting_room": " | ".join(meeting_rooms),
        "meeting_type": " | ".join(meeting_types),
        "openSection": section.get("openSection", ""),
        "linkIdentifier": section.get("linkIdentifier", ""),
        "isSectionLinked": section.get("isSectionLinked", ""),
        "sectionAttributes": attr_str,
    }


def scrape_all_courses(client, term_code, subjects, page_size=500, delay=0.15):
    """
    Scrape all course sections for the given term across all subjects.
    Returns a list of flat dicts.
    """
    all_rows = []

    for si, subj in enumerate(subjects):
        subj_code = subj["code"]
        subj_desc = subj["description"]
        offset = 0
        subject_count = 0

        # Reset and re-select term before each subject search
        client.reset_search()
        client.select_term(term_code)

        while True:
            try:
                data = client.search_courses(term_code, subj_code, offset, page_size)
            except Exception as e:
                log.error(f"Error fetching {subj_code} at offset {offset}: {e}")
                break

            if not data or "data" not in data or not data["data"]:
                # Log the raw response on first page if empty (debugging)
                if offset == 0:
                    log.debug(f"  Empty response for {subj_code}: {json.dumps(data)[:200]}")
                break

            sections = data["data"]
            total = data.get("totalCount", len(sections))

            for section in sections:
                row = extract_row(section)
                all_rows.append(row)
                subject_count += 1

            offset += page_size
            if offset >= total:
                break

            time.sleep(delay)

        status = f"({si + 1}/{len(subjects)})"
        print(f"  {status} {subj_desc} ({subj_code}): {subject_count} sections")
        if delay > 0:
            time.sleep(delay * 0.5)

    return all_rows


def save_csv(rows, path):
    """Save rows to a CSV file."""
    with open(path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=COLUMNS)
        writer.writeheader()
        writer.writerows(rows)
    log.info(f"Saved CSV: {path}")


def _coerce_int(value):
    """Convert Banner field values to integers when possible."""
    try:
        if value in ("", None):
            return 0
        return int(value)
    except (TypeError, ValueError):
        return 0


def build_web_bundle(rows, term_descriptions=None, generated_at=None):
    """
    Build a browser-friendly bundle that can be loaded from a local file.
    """
    term_descriptions = term_descriptions or {}
    sections = []
    term_codes = set()

    for row in rows:
        section = dict(row)
        term_code = str(section.get("term", "") or "")
        if term_code:
            term_codes.add(term_code)

        seats = _coerce_int(section.get("seatsAvailable"))
        wait_available = _coerce_int(section.get("waitAvailable"))
        if seats > 0:
            status = "open"
        elif wait_available > 0:
            status = "waitlist"
        else:
            status = "closed"

        section["status"] = status
        sections.append(section)

    terms = []
    for code in sorted(term_codes, reverse=True):
        terms.append({
            "code": code,
            "description": term_descriptions.get(code, f"Term {code}"),
        })

    return {
        "generatedAt": generated_at or datetime.now(timezone.utc).isoformat(),
        "terms": terms,
        "sections": sections,
    }


def save_web_bundle(rows, path, term_descriptions=None, generated_at=None):
    """Save a local JavaScript data bundle for index.html."""
    bundle = build_web_bundle(
        rows,
        term_descriptions=term_descriptions,
        generated_at=generated_at,
    )
    with open(path, "w", encoding="utf-8") as f:
        f.write("window.CNM_COURSE_DATA = ")
        json.dump(bundle, f, ensure_ascii=True, separators=(",", ":"))
        f.write(";\n")
    log.info(f"Saved web bundle: {path}")


def build_output_paths(output_dir, output_stem, term_code, bundle_name):
    """Build the default output paths for a scrape run."""
    return {
        "csv": output_dir / f"{output_stem}_{term_code}.csv",
        "xlsx": output_dir / f"{output_stem}_{term_code}.xlsx",
        "bundle": output_dir / bundle_name,
    }


def save_xlsx(rows, path, sheet_title="Courses"):
    """Save rows to an Excel file. Returns True on success, False if openpyxl missing."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        return False

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_title[:31]  # Excel sheet name max 31 chars

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")

    for col_idx, col_name in enumerate(COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for row_idx, row_data in enumerate(rows, 2):
        for col_idx, col_name in enumerate(COLUMNS, 1):
            val = row_data.get(col_name, "")
            if isinstance(val, str) and val.isdigit():
                val = int(val)
            ws.cell(row=row_idx, column=col_idx, value=val)

    for col_idx in range(1, len(COLUMNS) + 1):
        col_letter = get_column_letter(col_idx)
        max_len = len(COLUMNS[col_idx - 1])
        for row in ws.iter_rows(
            min_row=2, max_row=min(50, len(rows) + 1),
            min_col=col_idx, max_col=col_idx
        ):
            for cell in row:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 2, 40)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    wb.save(path)
    log.info(f"Saved Excel: {path}")
    return True


def run_cli(
    default_base_url=BASE_URL,
    default_school_label=DEFAULT_SCHOOL_LABEL,
    default_output_stem=DEFAULT_OUTPUT_STEM,
    default_bundle_name=DEFAULT_BUNDLE_NAME,
):
    parser = argparse.ArgumentParser(
        description=f"Scrape {default_school_label} course catalog from Banner"
    )
    parser.add_argument("--term", help="Term code (e.g. 202580). Default: first available.")
    parser.add_argument("--list-terms", action="store_true", help="List available terms and exit")
    parser.add_argument("--output-dir", default=".", help="Output directory (default: current)")
    parser.add_argument(
        "--base-url",
        default=default_base_url,
        help=f"Banner base URL (default: {default_base_url})",
    )
    parser.add_argument(
        "--school-label",
        default=default_school_label,
        help=f"Label used in console output (default: {default_school_label})",
    )
    parser.add_argument(
        "--output-stem",
        default=default_output_stem,
        help=f"Prefix for CSV/XLSX files (default: {default_output_stem})",
    )
    parser.add_argument(
        "--bundle-name",
        default=default_bundle_name,
        help=f"Filename for the JS bundle (default: {default_bundle_name})",
    )
    parser.add_argument("--debug", action="store_true", help="Enable debug logging")
    args = parser.parse_args()

    logging.basicConfig(
        level=logging.DEBUG if args.debug else logging.INFO,
        format="%(asctime)s [%(levelname)s] %(message)s",
    )

    output_dir = Path(args.output_dir)

    print("=" * 60)
    print(f"  {args.school_label} Course Scraper")
    print("=" * 60)
    print()

    # Step 1: Initialize
    print(f"[1/5] Connecting to {args.school_label} Banner...")
    client = BannerClient(base_url=args.base_url)
    client.initialize()
    print("  Session established.")

    # Step 2: Get terms
    print("[2/5] Fetching available terms...")
    terms = client.get_terms()
    if not terms:
        print("ERROR: No terms returned. The site may be down.")
        sys.exit(1)

    print(f"  Found {len(terms)} terms:")
    for i, t in enumerate(terms):
        print(f"    {i + 1}. {t['description']} (code: {t['code']})")

    if args.list_terms:
        sys.exit(0)

    # Pick term: use --term arg, or default to first (most recent)
    if args.term:
        term = next((t for t in terms if t["code"] == args.term), None)
        if not term:
            print(f"ERROR: Term '{args.term}' not found in available terms.")
            sys.exit(1)
    else:
        term = terms[0]

    term_code = term["code"]
    term_desc = term["description"]
    print(f"\n  Auto-selected: {term_desc} ({term_code})")

    # Step 3: Select term in server session (THE CRITICAL STEP)
    print(f"\n[3/5] Selecting term in Banner session...")
    client.select_term(term_code)
    print(f"  Term {term_code} locked into session.")

    # Step 4: Get subjects
    print(f"\n[4/5] Fetching subjects for {term_desc}...")
    subjects = client.get_subjects(term_code)
    print(f"  Found {len(subjects)} subjects.")

    # Step 5: Scrape all courses
    print(f"\n[5/5] Scraping courses...")
    all_rows = scrape_all_courses(client, term_code, subjects)

    print(f"\n  Total sections scraped: {len(all_rows)}")

    if not all_rows:
        print("  No data found. The term may not have courses listed yet.")
        sys.exit(0)

    # Save
    outputs = build_output_paths(output_dir, args.output_stem, term_code, args.bundle_name)
    csv_path = outputs["csv"]
    save_csv(all_rows, csv_path)
    print(f"\n  Saved CSV:   {csv_path}")

    bundle_path = outputs["bundle"]
    save_web_bundle(all_rows, bundle_path, term_descriptions={term_code: term_desc})
    print(f"  Saved Web:   {bundle_path}")

    xlsx_path = outputs["xlsx"]
    if save_xlsx(all_rows, xlsx_path, sheet_title=term_desc):
        print(f"  Saved Excel: {xlsx_path}")
    else:
        print("  (Install openpyxl for Excel output: pip install openpyxl)")

    print(f"\nDone! {len(all_rows)} sections saved.")


def main():
    run_cli()


if __name__ == "__main__":
    main()
